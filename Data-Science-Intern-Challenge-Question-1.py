import openpyxl

wb=openpyxl.load_workbook("2019 Winter Data Science Intern Challenge Data Set.xlsx")
sheet=wb["Sheet1"]
maxrow=sheet.max_row
maxlum=sheet.max_column
oasum=0
tisum=0
oa=0
ti=0
for i in range(1,maxlum+1):
    if sheet.cell(1,i).value=='order_amount':#get column of order_amount
        oa=i
    elif sheet.cell(1,i).value=='total_items':#get column of total_items
        ti=i
for i in range(2,maxrow+1):
    oasum=oasum+sheet.cell(i,oa).value #get the sum of order_amount
print(oasum)
for i in range(2,maxrow+1):
    tisum=tisum+sheet.cell(i,ti).value #get the sum of total_items
print(tisum)
aov=round(oasum/tisum,2)
print(aov)#output
